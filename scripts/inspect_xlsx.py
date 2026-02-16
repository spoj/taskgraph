#!/usr/bin/env python3
"""Inspect xlsx files iteratively.

Usage:
  inspect_xlsx.py <file>                      # list sheets with dimensions
  inspect_xlsx.py <file> <sheet>              # preview first 50 rows (grid)
  inspect_xlsx.py <file> <sheet> <range>      # show cell range as grid
  inspect_xlsx.py <file> <sheet> -d           # detail: cell / formula / value
  inspect_xlsx.py <file> <sheet> <range> -d   # detail for specific range

Range formats: A1:G20, 1:50 (row range, all cols), A:G (col range, all rows)
Modes: default is grid (values only), -d for detail (location + formula + value)
"""

import re
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string, get_column_letter
except ModuleNotFoundError as e:
    raise SystemExit(
        "openpyxl is required for inspect_xlsx.py. Install it with: uv add openpyxl"
    ) from e


def _col(s: str) -> int:
    """Column letter(s) -> 1-based index."""
    return column_index_from_string(s.upper())


def _parse_range(ws: Any, range_str: str) -> tuple[int, int, int, int]:
    """Parse range string into (min_row, max_row, min_col, max_col)."""
    range_str = range_str.strip()

    # Row-only range: "1:50"
    m = re.fullmatch(r"(\d+):(\d+)", range_str)
    if m:
        return int(m[1]), int(m[2]), 1, ws.max_column

    # Col-only range: "A:G"
    m = re.fullmatch(r"([A-Za-z]+):([A-Za-z]+)", range_str)
    if m:
        return 1, ws.max_row, _col(m[1]), _col(m[2])

    # Full range: "A1:G20"
    m = re.fullmatch(r"([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)", range_str)
    if m:
        return int(m[2]), int(m[4]), _col(m[1]), _col(m[3])

    # Single cell: "B5"
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", range_str)
    if m:
        r, c = int(m[2]), _col(m[1])
        return r, r, c, c

    print(f"Error: cannot parse range '{range_str}'")
    print("Expected: A1:G20, 1:50, A:G, or B5")
    sys.exit(1)


def _fmt(value: object | None, width: int = 60) -> str:
    if value is None:
        return ""
    return str(value)[:width]


def list_sheets(path: Path) -> None:
    """Mode 1: overview of all sheets."""
    wb = openpyxl.load_workbook(path, data_only=True)
    size = path.stat().st_size
    print(f"\n{path.name}  ({size:,} bytes)")
    print(f"{'Sheet':<30} {'Rows':>8} {'Cols':>8}  Col range")
    print("-" * 70)
    for name in wb.sheetnames:
        ws = wb[name]
        max_col = ws.max_column or 0
        max_row = ws.max_row or 0
        col_range = f"A:{get_column_letter(max_col)}" if max_col else "-"
        print(f"{name:<30} {max_row:>8} {max_col:>8}  {col_range}")
    print()


def _resolve_bounds(
    ws: Any, range_str: str | None, max_rows: int = 50
) -> tuple[int, int, int, int, bool]:
    """Return (min_row, max_row, min_col, max_col, clamped) for sheet/range."""
    if range_str:
        min_row, max_row, min_col, max_col = _parse_range(ws, range_str)
        max_row = min(max_row, ws.max_row or 0)
        max_col = min(max_col, ws.max_column or 0)
        return min_row, max_row, min_col, max_col, False
    else:
        total = ws.max_row or 0
        show = min(max_rows, total)
        return 1, show, 1, ws.max_column or 0, total > show


def _open_sheet(path: Path, sheet: str, data_only: bool = True) -> tuple[Any, Any]:
    """Open workbook and return (wb, ws), exiting on bad sheet name."""
    wb = openpyxl.load_workbook(path, data_only=data_only)
    if sheet not in wb.sheetnames:
        print(f"Error: sheet '{sheet}' not found. Available: {wb.sheetnames}")
        sys.exit(1)
    return wb, wb[sheet]


def _print_header(
    sheet: str,
    ws: Any,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    """Print range header line."""
    c0 = get_column_letter(min_col)
    c1 = get_column_letter(max_col)
    total = ws.max_row or 0
    total_cols = ws.max_column or 0
    print(
        f"\n{sheet}  {c0}{min_row}:{c1}{max_row}"
        f"  ({max_row - min_row + 1} rows x {max_col - min_col + 1} cols"
        f", sheet has {total} rows x {total_cols} cols)"
    )
    print("-" * 70)


# ---------------------------------------------------------------------------
# Grid mode (default): compact value grid
# ---------------------------------------------------------------------------


def show_grid(path: Path, sheet: str, range_str: str | None = None) -> None:
    """Show values in a compact grid."""
    _, ws = _open_sheet(path, sheet)
    min_row, max_row, min_col, max_col, clamped = _resolve_bounds(ws, range_str)
    _print_header(sheet, ws, min_row, max_row, min_col, max_col)
    _print_rows_grid(ws, min_row, max_row, min_col, max_col)
    if clamped:
        total = ws.max_row or 0
        print(
            f"  ... {total - max_row} more rows (use range to see more, e.g. 1:{total})"
        )
    print()


def _print_rows_grid(
    ws: Any, min_row: int, max_row: int, min_col: int, max_col: int
) -> None:
    row_label_w = len(str(max_row)) + 1
    letters = [f"{get_column_letter(c):>12}" for c in range(min_col, max_col + 1)]
    print(f"{'':>{row_label_w}}  {''.join(letters)}")

    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=False,
    ):
        rnum = row[0].row
        vals = [f"{_fmt(c.value, 12):>12}" for c in row]
        line = "".join(vals)
        if line.strip():
            print(f"{rnum:>{row_label_w}}  {line}")


# ---------------------------------------------------------------------------
# Detail mode (-d): long table with Cell | Formula | Value
# ---------------------------------------------------------------------------


def show_detail(path: Path, sheet: str, range_str: str | None = None) -> None:
    """Show cell-by-cell detail: location, formula, computed value."""
    # Need two loads: one for formulas, one for computed values
    _, ws_formula = _open_sheet(path, sheet, data_only=False)
    _, ws_value = _open_sheet(path, sheet, data_only=True)

    min_row, max_row, min_col, max_col, clamped = _resolve_bounds(ws_formula, range_str)
    _print_header(sheet, ws_formula, min_row, max_row, min_col, max_col)

    print(f"{'Cell':<8} {'Formula':<40} {'Value'}")
    print(f"{'----':<8} {'-------':<40} {'-----'}")

    for row_f, row_v in zip(
        ws_formula.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=False,
        ),
        ws_value.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=False,
        ),
    ):
        for cf, cv in zip(row_f, row_v):
            formula_raw = cf.value
            value = cv.value
            if formula_raw is None and value is None:
                continue
            coord = f"{get_column_letter(cf.column or 1)}{cf.row}"
            # Show formula if it starts with =, otherwise show raw content
            if isinstance(formula_raw, str) and formula_raw.startswith("="):
                formula_str = formula_raw[:40]
            else:
                formula_str = ""
            val_str = _fmt(value, 60)
            print(f"{coord:<8} {formula_str:<40} {val_str}")

    if clamped:
        total = ws_formula.max_row or 0
        print(f"  ... {total - max_row} more rows")
    print()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    flags = {a for a in sys.argv[1:] if a.startswith("-")}
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    detail = "-d" in flags

    if not args:
        print(__doc__)
        sys.exit(1)

    path = Path(args[0])
    if not path.exists():
        print(f"Error: file not found: {path}")
        sys.exit(1)

    if len(args) == 1:
        list_sheets(path)
    elif len(args) == 2:
        if detail:
            show_detail(path, args[1])
        else:
            show_grid(path, args[1])
    elif len(args) >= 3:
        if detail:
            show_detail(path, args[1], args[2])
        else:
            show_grid(path, args[1], args[2])
    else:
        print(__doc__)
        sys.exit(1)


if __name__ == "__main__":
    main()
