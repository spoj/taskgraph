import asyncio
from pathlib import Path

import pytest
import polars as pl
from openpyxl import Workbook

from tests.conftest import _make_node, _views, _tables
from src.ingest import (
    coerce_to_dataframe,
    ingest_csv,
    ingest_excel,
    ingest_parquet,
    ingest_pdf,
    ingest_llm_source,
    ingest_llm_pages,
    ingest_table,
    llm,
    llm_pages,
    FileInput,
    LLMSource,
    LLMPagesSource,
    PDF_EXTRACTION_PROMPT,
    PDF_PAGES_EXTRACTION_PROMPT,
    PDF_RESPONSE_FORMAT,
    parse_file_path,
    parse_file_string,
    _split_pdf_pages,
    _rows_match,
    _extract_page_with_majority,
)
from src.task import Node
from src.agent import run_validate_sql, validate_node_complete
from src.workspace import Workspace, materialize_node_outputs


class TestIngestion:
    """Tests for data ingestion: coerce_to_dataframe, ingest_table."""

    def test_coerce_dataframe(self):
        """DataFrame passes through unchanged."""
        df = pl.DataFrame({"a": [1, 2], "b": [3, 4]})
        result = coerce_to_dataframe(df)
        assert isinstance(result, pl.DataFrame)
        assert result.columns == ["a", "b"]
        assert len(result) == 2

    def test_coerce_list_of_dicts(self):
        """list[dict] is converted to DataFrame."""
        data = [{"x": 1, "y": "a"}, {"x": 2, "y": "b"}]
        result = coerce_to_dataframe(data)
        assert isinstance(result, pl.DataFrame)
        assert result.columns == ["x", "y"]
        assert len(result) == 2

    def test_coerce_dict_of_lists(self):
        """dict[str, list] is converted to DataFrame."""
        data = {"col1": [10, 20, 30], "col2": ["a", "b", "c"]}
        result = coerce_to_dataframe(data)
        assert isinstance(result, pl.DataFrame)
        assert result.columns == ["col1", "col2"]
        assert len(result) == 3

    def test_coerce_unsupported_type(self):
        """Unsupported types raise TypeError."""
        with pytest.raises(TypeError, match="Unsupported"):
            coerce_to_dataframe("not a table")
        with pytest.raises(TypeError, match="Unsupported"):
            coerce_to_dataframe(42)

    def test_ingest_list_of_dicts(self, conn):
        """Ingest list[dict] creates table with _row_id."""
        data = [{"name": "alice", "age": 30}, {"name": "bob", "age": 25}]
        ingest_table(conn, data, "people")

        rows = conn.execute(
            "SELECT _row_id, name, age FROM people ORDER BY _row_id"
        ).fetchall()
        assert rows == [(1, "alice", 30), (2, "bob", 25)]

    def test_ingest_dataframe(self, conn):
        """Ingest DataFrame creates table with _row_id."""
        df = pl.DataFrame({"product": ["A", "B"], "price": [10.0, 20.0]})
        ingest_table(conn, df, "products")

        rows = conn.execute(
            "SELECT _row_id, product, price FROM products ORDER BY _row_id"
        ).fetchall()
        assert rows == [(1, "A", 10.0), (2, "B", 20.0)]

    def test_ingest_dict_of_lists(self, conn):
        """Ingest dict[str, list] creates table with _row_id."""
        data = {"city": ["NYC", "LA", "CHI"], "pop": [8, 4, 3]}
        ingest_table(conn, data, "cities")

        count = conn.execute("SELECT COUNT(*) FROM cities").fetchone()[0]
        assert count == 3

        # Check _row_id exists and is sequential
        ids = conn.execute("SELECT _row_id FROM cities ORDER BY _row_id").fetchall()
        assert ids == [(1,), (2,), (3,)]

    def test_row_id_is_integer(self, conn):
        """_row_id column is INTEGER type."""
        ingest_table(conn, [{"x": 1}], "t")
        cols = conn.execute(
            "SELECT column_name, data_type FROM information_schema.columns "
            "WHERE table_name = 't' AND column_name = '_row_id'"
        ).fetchall()
        assert len(cols) == 1
        assert cols[0][1] == "INTEGER"

    def test_ingest_overwrites_existing(self, conn):
        """Ingesting to an existing table name drops and recreates it."""
        ingest_table(conn, [{"v": 1}], "t")
        assert conn.execute("SELECT COUNT(*) FROM t").fetchone()[0] == 1

        ingest_table(conn, [{"v": 10}, {"v": 20}], "t")
        assert conn.execute("SELECT COUNT(*) FROM t").fetchone()[0] == 2

    def test_ingest_preserves_types(self, conn):
        """Ingested data preserves column types."""
        data = [
            {"int_col": 42, "float_col": 3.14, "str_col": "hello", "bool_col": True},
        ]
        ingest_table(conn, data, "typed")

        row = conn.execute(
            "SELECT int_col, float_col, str_col, bool_col FROM typed"
        ).fetchone()
        assert row[0] == 42
        assert abs(row[1] - 3.14) < 0.001
        assert row[2] == "hello"
        assert row[3] is True

    def test_ingest_empty_list(self, conn):
        """Ingesting empty list creates table with zero rows."""
        df = pl.DataFrame({"x": pl.Series([], dtype=pl.Int64)})
        ingest_table(conn, df, "empty_t")
        count = conn.execute("SELECT COUNT(*) FROM empty_t").fetchone()[0]
        assert count == 0

    def test_ingest_csv_file(self, conn, tmp_path: Path):
        df = pl.DataFrame({"id": [1, 2], "val": ["a", "b"]})
        path = tmp_path / "data.csv"
        df.write_csv(path)

        ingest_csv(conn, path, "csv_t")
        rows = conn.execute(
            "SELECT _row_id, id, val FROM csv_t ORDER BY _row_id"
        ).fetchall()
        assert rows == [(1, 1, "a"), (2, 2, "b")]

    def test_ingest_parquet_file(self, conn, tmp_path: Path):
        df = pl.DataFrame({"id": [1, 2], "val": ["a", "b"]})
        path = tmp_path / "data.parquet"
        df.write_parquet(path)

        ingest_parquet(conn, path, "parquet_t")
        rows = conn.execute(
            "SELECT _row_id, id, val FROM parquet_t ORDER BY _row_id"
        ).fetchall()
        assert rows == [(1, 1, "a"), (2, 2, "b")]

    def test_ingest_excel_file(self, conn, tmp_path: Path):
        path = tmp_path / "data.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["id", "val"])
        ws.append([1, "a"])
        ws.append([2, "b"])
        wb.save(path)

        ingest_excel(conn, path, "excel_t")
        # header=false: columns are A1, B1; header row becomes data
        rows = conn.execute(
            "SELECT _row_id, A1, B1 FROM excel_t ORDER BY _row_id"
        ).fetchall()
        assert rows == [(1, "id", "val"), (2, "1", "a"), (3, "2", "b")]

    def test_ingest_excel_specific_sheet(self, conn, tmp_path: Path):
        path = tmp_path / "data.xlsx"
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["id", "val"])
        ws1.append([1, "a"])
        ws2 = wb.create_sheet(title="Budget")
        ws2.append(["id", "val"])
        ws2.append([99, "z"])
        wb.save(path)

        ingest_excel(conn, path, "excel_t", sheet="Budget")
        # header=false: columns are A1, B1; header row becomes data
        rows = conn.execute(
            "SELECT _row_id, A1, B1 FROM excel_t ORDER BY _row_id"
        ).fetchall()
        assert rows == [(1, "id", "val"), (2, "99", "z")]

    def test_ingest_pdf_file(self, conn, tmp_path: Path):
        path = tmp_path / "data.pdf"
        path.write_bytes(b"%PDF-1.4\n%fake\n")

        class DummyClient:
            def __init__(self):
                self.calls = []

            async def chat(self, model, messages, response_format=None):
                self.calls.append(
                    {
                        "model": model,
                        "messages": messages,
                        "response_format": response_format,
                    }
                )
                return {
                    "message": {
                        "content": '{"rows": [{"id": 1, "val": "a"}, {"id": 2, "val": "b"}]}'
                    },
                    "usage": {},
                }

        client = DummyClient()
        asyncio.run(ingest_pdf(conn, path, "pdf_t", client=client))

        rows = conn.execute(
            "SELECT _row_id, id, val FROM pdf_t ORDER BY _row_id"
        ).fetchall()
        assert rows == [(1, 1, "a"), (2, 2, "b")]
        assert client.calls
        content = client.calls[0]["messages"][0]["content"]
        assert any(item.get("type") == "image_url" for item in content)
        assert client.calls[0]["response_format"] == PDF_RESPONSE_FORMAT

    def test_parse_file_string(self, tmp_path: Path):
        base_dir = tmp_path
        file_input = parse_file_string("data.xlsx#Sheet2", base_dir=base_dir)
        assert file_input.format == "excel"
        assert file_input.sheet == "Sheet2"
        assert str(file_input.path).endswith("data.xlsx")

    def test_parse_file_path(self, tmp_path: Path):
        path = tmp_path / "data.csv"
        file_input = parse_file_path(path, base_dir=tmp_path)
        assert file_input.format == "csv"
        assert file_input.sheet is None


class TestSourceNodeValidation:
    """Tests for source node validation via Node.validate_outputs() and
    the unified post-execution flow (validate_node_complete + materialize).

    In the unified node model, source nodes use:
    - ``columns`` for required column checks (via Node.validate_outputs)
    - ``validate`` queries for validation views (via run_validate_sql + validate_validation_views)
    - ``validate_node_complete()`` + ``materialize_node_outputs()`` for the full flow
    """

    def test_source_columns_pass(self, conn):
        """Source column validation passes when all required columns exist."""
        ingest_table(conn, [{"name": "alice", "age": 30}], "people")
        node = Node(
            name="people",
            source=[{"name": "alice", "age": 30}],
            columns=["name", "age"],
        )
        errors = node.validate_outputs(conn)
        assert errors == []

    def test_source_columns_missing(self, conn):
        """Source column validation fails when required columns are missing."""
        ingest_table(conn, [{"name": "alice"}], "people")
        node = Node(
            name="people", source=[{"name": "alice"}], columns=["name", "age", "email"]
        )
        errors = node.validate_outputs(conn)
        assert len(errors) == 1
        assert "age" in errors[0]
        assert "email" in errors[0]
        assert "name" in errors[0]  # actual columns listed

    def test_source_columns_table_not_found(self, conn):
        """Source column validation reports missing table."""
        node = Node(name="nonexistent", source=[], columns=["col1"])
        errors = node.validate_outputs(conn)
        assert len(errors) == 1
        assert "was not created" in errors[0].lower()

    def test_source_columns_extra_columns_ok(self, conn):
        """Extra columns beyond required are fine."""
        ingest_table(conn, [{"a": 1, "b": 2, "c": 3}], "t")
        node = Node(name="t", source=[], columns=["a"])
        errors = node.validate_outputs(conn)
        assert errors == []

    def test_source_columns_excludes_row_id(self, conn):
        """_row_id is excluded from the actual columns listed in error messages."""
        ingest_table(conn, [{"x": 1}], "t")
        node = Node(name="t", source=[], columns=["missing_col"])
        errors = node.validate_outputs(conn)
        assert len(errors) == 1
        assert "_row_id" not in errors[0]

    def test_source_columns_short_circuits_before_validation(self, conn):
        """Column errors are found before validation would run."""
        ingest_table(conn, [{"x": 1}], "t")
        node = Node(
            name="t",
            source=[],
            columns=["missing"],
            validate={"main": "SELECT 'fail' AS status, 'should not run' AS message"},
        )
        # validate_outputs catches the column error first
        errors = node.validate_outputs(conn)
        assert len(errors) == 1
        assert "missing" in errors[0]

    def test_validate_sql_pass(self, conn):
        """Validation passes when all rows have status='pass'."""
        ingest_table(conn, [{"id": 1, "val": 10}, {"id": 2, "val": 20}], "data")
        node = Node(
            name="data",
            source=[],
            validate={
                "main": "SELECT 'pass' AS status, 'all values positive' AS message"
            },
        )
        errors = run_validate_sql(conn=conn, node=node)
        assert errors == []
        errors = node.validate_validation_views(conn)
        assert errors == []

    def test_validate_sql_fail(self, conn):
        """Validation fails when view has status='fail' rows."""
        ingest_table(conn, [{"id": 1, "val": -5}], "data")
        node = Node(
            name="data",
            source=[],
            validate={
                "main": (
                    "SELECT 'fail' AS status, "
                    "'negative value for id=' || CAST(id AS VARCHAR) AS message "
                    "FROM data WHERE val < 0"
                )
            },
        )
        errors = run_validate_sql(conn=conn, node=node)
        assert errors == []  # SQL execution succeeds
        errors = node.validate_validation_views(conn)
        assert len(errors) == 1
        assert "negative value for id=1" in errors[0]

    def test_validate_sql_error_handling(self, conn):
        """SQL validation catches query errors gracefully."""
        node = Node(
            name="bad",
            source=[],
            validate={
                "main": ("SELECT 'fail' AS status, x AS message FROM nonexistent_table")
            },
        )
        errors = run_validate_sql(conn=conn, node=node)
        assert len(errors) == 1
        assert "error" in errors[0].lower() or "nonexistent_table" in errors[0].lower()

    def test_validate_sql_mixed_status(self, conn):
        """Validation view with mixed pass/fail rows reports only failures."""
        ingest_table(
            conn, [{"id": 1, "status": "bad"}, {"id": 2, "status": "ok"}], "items"
        )
        node = Node(
            name="items",
            source=[],
            validate={
                "main": (
                    "SELECT 'fail' AS status, "
                    "'bad status for id=' || CAST(id AS VARCHAR) AS message "
                    "FROM items WHERE status = 'bad' "
                    "UNION ALL "
                    "SELECT 'pass' AS status, 'ok' AS message "
                    "FROM items WHERE status = 'ok'"
                )
            },
        )
        errors = run_validate_sql(conn=conn, node=node)
        assert errors == []
        errors = node.validate_validation_views(conn)
        assert len(errors) == 1
        assert "bad status for id=1" in errors[0]

    def test_validate_sql_multiple_views_both_fail(self, conn):
        """Multiple validation views both reporting failures."""
        ingest_table(conn, [{"id": 1}], "t")
        node = Node(
            name="t",
            source=[],
            validate={
                "main": "SELECT 'fail' AS status, 'error1' AS message",
                "extra": "SELECT 'fail' AS status, 'error2' AS message",
            },
        )
        errors = run_validate_sql(conn=conn, node=node)
        assert errors == []
        errors = node.validate_validation_views(conn)
        assert len(errors) >= 1
        all_err = "\n".join(errors)
        assert "error1" in all_err
        assert "error2" in all_err

    def test_post_execute_materializes_validation_views(self, conn):
        """Validation views are materialized as tables after passing."""
        ingest_table(conn, [{"id": 1}], "t")
        node = Node(
            name="t",
            source=[],
            validate={"main": "SELECT 'pass' AS status, 'ok' AS message"},
        )
        # Run validation (creates validation views), then materialize
        errors = validate_node_complete(conn, node)
        assert not errors
        materialize_node_outputs(conn, node)
        # View should be gone (materialized into a table)
        assert "t__validation_main" not in _views(conn)
        # Table should exist with same data
        assert "t__validation_main" in _tables(conn)
        row = conn.execute("SELECT status, message FROM t__validation_main").fetchone()
        assert row == ("pass", "ok")

    def test_validate_sql_preserves_definition(self, conn):
        """Validation SQL is recorded in _trace (visible via _view_definitions)."""
        ingest_table(conn, [{"id": 1}], "data")
        node = Node(
            name="data",
            source=[],
            validate={"main": "SELECT 'pass' AS status, 'all good' AS message"},
        )
        errors = run_validate_sql(conn=conn, node=node)
        assert errors == []
        # SQL definition should be stored with node name as task label
        rows = conn.execute(
            "SELECT node, view_name, sql FROM _view_definitions"
        ).fetchall()
        assert len(rows) == 1
        assert rows[0][0] == "data"  # label = node name
        assert rows[0][1] == "data__validation_main"
        assert "all good" in rows[0][2]

    def test_post_execute_multiple_validation_views(self, conn):
        """Multiple validation views are all materialized."""
        ingest_table(conn, [{"id": 1}], "t")
        node = Node(
            name="t",
            source=[],
            validate={
                "main": "SELECT 'pass' AS status, 'ok' AS message",
                "extra": "SELECT 'pass' AS status, 'also ok' AS message",
            },
        )
        errors = validate_node_complete(conn, node)
        assert not errors
        materialize_node_outputs(conn, node)
        assert {"t__validation_main", "t__validation_extra"} <= _tables(conn)

    def test_no_validation_returns_empty(self, conn):
        """Source node with no columns or validate passes validation."""
        ingest_table(conn, [{"x": 1}], "t")
        node = Node(name="t", source=[])
        errors = node.validate_outputs(conn)
        assert errors == []

    def test_multiple_source_nodes_column_check(self, conn):
        """Column validation works across multiple source nodes."""
        ingest_table(conn, [{"a": 1, "b": 2}], "t1")
        ingest_table(conn, [{"x": 1}], "t2")
        node1 = Node(name="t1", source=[], columns=["a", "b"])
        node2 = Node(name="t2", source=[], columns=["x", "y"])  # y is missing
        assert node1.validate_outputs(conn) == []
        errors = node2.validate_outputs(conn)
        assert len(errors) == 1
        assert "t2" in errors[0]
        assert "y" in errors[0]

    def test_source_callable_error_handling(self):
        """Source node callable errors are caught with context."""

        def bad_loader():
            raise FileNotFoundError("data.csv not found")

        node = Node(name="broken", source=bad_loader)
        ws = Workspace(db_path=":memory:", nodes=[node])
        import duckdb

        conn = duckdb.connect(":memory:")
        try:
            with pytest.raises(RuntimeError, match="Source 'broken' callable failed"):
                asyncio.run(ws._ingest_node(conn, node, client=None))
        finally:
            conn.close()


class TestLLMSource:
    """Tests for llm() factory and multi-file LLM ingestion."""

    def test_llm_factory_basic(self):
        """llm() creates LLMSource with default prompt."""
        src = llm("a.pdf", "b.pdf")
        assert isinstance(src, LLMSource)
        assert src.files == ("a.pdf", "b.pdf")
        assert src.prompt == PDF_EXTRACTION_PROMPT

    def test_llm_factory_custom_prompt(self):
        """llm() accepts a custom prompt."""
        src = llm("a.pdf", prompt="Extract invoices")
        assert src.prompt == "Extract invoices"
        assert src.files == ("a.pdf",)

    def test_llm_factory_no_files(self):
        """llm() with no files raises ValueError."""
        with pytest.raises(ValueError, match="at least one file"):
            llm()

    def test_llm_factory_non_string_file(self):
        """llm() with non-string file raises TypeError."""
        with pytest.raises(TypeError, match="must be strings"):
            llm(123)

    def test_ingest_llm_source_single_file(self, conn, tmp_path: Path):
        """LLMSource with a single PDF extracts and ingests rows."""
        path = tmp_path / "doc.pdf"
        path.write_bytes(b"%PDF-1.4\n%fake\n")

        class DummyClient:
            async def chat(self, model, messages, response_format=None):
                return {
                    "message": {"content": '{"rows": [{"id": 1, "val": "x"}]}'},
                    "usage": {},
                }

        source = LLMSource(
            files=(FileInput(path=path, format="pdf"),),
            prompt=PDF_EXTRACTION_PROMPT,
        )
        asyncio.run(ingest_llm_source(conn, source, "llm_t", client=DummyClient()))
        rows = conn.execute(
            "SELECT _row_id, id, val FROM llm_t ORDER BY _row_id"
        ).fetchall()
        assert rows == [(1, 1, "x")]

    def test_ingest_llm_source_multiple_files(self, conn, tmp_path: Path):
        """LLMSource with multiple PDFs concatenates rows into one table."""
        path_a = tmp_path / "a.pdf"
        path_b = tmp_path / "b.pdf"
        path_a.write_bytes(b"%PDF-1.4\n%fake a\n")
        path_b.write_bytes(b"%PDF-1.4\n%fake b\n")

        call_count = 0

        class DummyClient:
            async def chat(self, model, messages, response_format=None):
                nonlocal call_count
                call_count += 1
                if call_count == 1:
                    return {
                        "message": {"content": '{"rows": [{"id": 1, "val": "a"}]}'},
                        "usage": {},
                    }
                return {
                    "message": {"content": '{"rows": [{"id": 2, "val": "b"}]}'},
                    "usage": {},
                }

        source = LLMSource(
            files=(
                FileInput(path=path_a, format="pdf"),
                FileInput(path=path_b, format="pdf"),
            ),
            prompt=PDF_EXTRACTION_PROMPT,
        )
        asyncio.run(ingest_llm_source(conn, source, "multi_t", client=DummyClient()))
        rows = conn.execute(
            "SELECT _row_id, id, val FROM multi_t ORDER BY _row_id"
        ).fetchall()
        assert rows == [(1, 1, "a"), (2, 2, "b")]
        assert call_count == 2

    def test_ingest_llm_source_custom_prompt(self, conn, tmp_path: Path):
        """LLMSource passes custom prompt to the LLM."""
        path = tmp_path / "doc.pdf"
        path.write_bytes(b"%PDF-1.4\n%fake\n")
        captured_prompt = []

        class DummyClient:
            async def chat(self, model, messages, response_format=None):
                text_parts = [
                    c["text"] for c in messages[0]["content"] if c.get("type") == "text"
                ]
                captured_prompt.extend(text_parts)
                return {
                    "message": {"content": '{"rows": [{"x": 1}]}'},
                    "usage": {},
                }

        source = LLMSource(
            files=(FileInput(path=path, format="pdf"),),
            prompt="Extract all line items",
        )
        asyncio.run(ingest_llm_source(conn, source, "cust_t", client=DummyClient()))
        assert captured_prompt == ["Extract all line items"]

    def test_ingest_llm_source_no_client(self, conn):
        """LLMSource without client raises RuntimeError."""
        source = LLMSource(files=(FileInput(path=Path("x.pdf"), format="pdf"),))
        with pytest.raises(RuntimeError, match="requires an OpenRouterClient"):
            asyncio.run(ingest_llm_source(conn, source, "t", client=None))


def _make_pdf(num_pages: int = 1) -> bytes:
    """Create a minimal valid multi-page PDF using pypdf."""
    from pypdf import PdfWriter

    writer = PdfWriter()
    for _ in range(num_pages):
        writer.add_blank_page(width=612, height=792)
    buf = __import__("io").BytesIO()
    writer.write(buf)
    return buf.getvalue()


class TestLLMPagesSource:
    """Tests for llm_pages() factory, page splitting, majority vote, and full pipeline."""

    # -- factory tests --

    def test_llm_pages_factory_basic(self):
        """llm_pages() creates LLMPagesSource with default prompt."""
        src = llm_pages("a.pdf", "b.pdf")
        assert isinstance(src, LLMPagesSource)
        assert src.files == ("a.pdf", "b.pdf")
        assert src.prompt == PDF_PAGES_EXTRACTION_PROMPT

    def test_llm_pages_factory_custom_prompt(self):
        """llm_pages() accepts a custom prompt."""
        src = llm_pages("a.pdf", prompt="Extract transactions")
        assert src.prompt == "Extract transactions"

    def test_llm_pages_factory_no_files(self):
        with pytest.raises(ValueError, match="at least one file"):
            llm_pages()

    def test_llm_pages_factory_non_string(self):
        with pytest.raises(TypeError, match="must be strings"):
            llm_pages(42)

    # -- page splitting --

    def test_split_pdf_pages_single(self, tmp_path: Path):
        """Splitting a 1-page PDF returns 1 page."""
        path = tmp_path / "one.pdf"
        path.write_bytes(_make_pdf(1))
        pages = _split_pdf_pages(path)
        assert len(pages) == 1
        assert isinstance(pages[0], bytes)

    def test_split_pdf_pages_multi(self, tmp_path: Path):
        """Splitting a 3-page PDF returns 3 pages."""
        path = tmp_path / "three.pdf"
        path.write_bytes(_make_pdf(3))
        pages = _split_pdf_pages(path)
        assert len(pages) == 3

    def test_split_pdf_pages_file_not_found(self, tmp_path: Path):
        with pytest.raises(FileNotFoundError):
            _split_pdf_pages(tmp_path / "nope.pdf")

    def test_split_pdf_pages_empty_file(self, tmp_path: Path):
        path = tmp_path / "empty.pdf"
        path.write_bytes(b"")
        with pytest.raises(RuntimeError, match="empty"):
            _split_pdf_pages(path)

    # -- rows_match --

    def test_rows_match_identical(self):
        a = [{"id": 1, "val": "x"}]
        assert _rows_match(a, a)

    def test_rows_match_whitespace(self):
        """Trailing/leading whitespace is ignored."""
        a = [{"id": 1, "val": "hello"}]
        b = [{"id": 1, "val": "  hello  "}]
        assert _rows_match(a, b)

    def test_rows_match_different(self):
        a = [{"id": 1}]
        b = [{"id": 2}]
        assert not _rows_match(a, b)

    def test_rows_match_different_length(self):
        a = [{"id": 1}]
        b = [{"id": 1}, {"id": 2}]
        assert not _rows_match(a, b)

    def test_rows_match_both_empty(self):
        assert _rows_match([], [])

    # -- majority vote --

    def test_majority_two_passes_agree(self, tmp_path: Path):
        """If pass 1 and 2 agree, only 2 LLM calls are made."""
        page_pdf = _make_pdf(1)
        call_count = 0

        class Client:
            async def chat(self, model, messages, response_format=None):
                nonlocal call_count
                call_count += 1
                return {
                    "message": {"content": '{"rows": [{"id": 1}]}'},
                    "usage": {},
                }

        rows = asyncio.run(
            _extract_page_with_majority(Client(), page_pdf, "Extract", "=== p1 ===")
        )
        assert rows == [{"id": 1}]
        assert call_count == 2

    def test_majority_third_pass_tiebreaker(self, tmp_path: Path):
        """If passes 1 and 2 disagree but 3 matches one of them, majority wins."""
        page_pdf = _make_pdf(1)
        call_count = 0

        class Client:
            async def chat(self, model, messages, response_format=None):
                nonlocal call_count
                call_count += 1
                if call_count == 1:
                    return {
                        "message": {"content": '{"rows": [{"id": 1}]}'},
                        "usage": {},
                    }
                # passes 2 and 3 agree on id=2
                return {
                    "message": {"content": '{"rows": [{"id": 2}]}'},
                    "usage": {},
                }

        rows = asyncio.run(
            _extract_page_with_majority(Client(), page_pdf, "Extract", "=== p1 ===")
        )
        assert rows == [{"id": 2}]
        assert call_count == 3

    def test_majority_all_three_differ_errors(self, tmp_path: Path):
        """If all 3 passes produce different results, raises RuntimeError."""
        page_pdf = _make_pdf(1)
        call_count = 0

        class Client:
            async def chat(self, model, messages, response_format=None):
                nonlocal call_count
                call_count += 1
                return {
                    "message": {"content": f'{{"rows": [{{"id": {call_count}}}]}}'},
                    "usage": {},
                }

        with pytest.raises(RuntimeError, match="All 3 extraction passes"):
            asyncio.run(
                _extract_page_with_majority(Client(), page_pdf, "Extract", "=== p1 ===")
            )
        assert call_count == 3

    # -- full pipeline --

    def test_ingest_llm_pages_single_page(self, conn, tmp_path: Path):
        """Single-page PDF: 2 matching passes → 1 table."""
        path = tmp_path / "stmt.pdf"
        path.write_bytes(_make_pdf(1))

        class Client:
            async def chat(self, model, messages, response_format=None):
                return {
                    "message": {
                        "content": '{"rows": [{"date": "2025-01", "amt": 100}]}'
                    },
                    "usage": {},
                }

        source = LLMPagesSource(
            files=(FileInput(path=path, format="pdf"),),
        )
        asyncio.run(ingest_llm_pages(conn, source, "pages_t", client=Client()))
        rows = conn.execute(
            "SELECT _row_id, date, amt FROM pages_t ORDER BY _row_id"
        ).fetchall()
        assert rows == [(1, "2025-01", 100)]

    def test_ingest_llm_pages_multi_page(self, conn, tmp_path: Path):
        """2-page PDF: each page produces different rows, all concatenated."""
        path = tmp_path / "stmt.pdf"
        path.write_bytes(_make_pdf(2))
        call_count = 0

        class Client:
            async def chat(self, model, messages, response_format=None):
                nonlocal call_count
                call_count += 1
                # calls 1-2 are page 1 (2 passes), calls 3-4 are page 2
                page_num = (call_count - 1) // 2 + 1
                return {
                    "message": {
                        "content": f'{{"rows": [{{"page": {page_num}, "val": "r{page_num}"}}]}}'
                    },
                    "usage": {},
                }

        source = LLMPagesSource(
            files=(FileInput(path=path, format="pdf"),),
        )
        asyncio.run(ingest_llm_pages(conn, source, "mp_t", client=Client()))
        rows = conn.execute(
            "SELECT _row_id, page, val FROM mp_t ORDER BY _row_id"
        ).fetchall()
        assert rows == [(1, 1, "r1"), (2, 2, "r2")]
        assert call_count == 4  # 2 passes * 2 pages

    def test_ingest_llm_pages_multiple_files(self, conn, tmp_path: Path):
        """Two 1-page PDFs: rows from both files concatenated."""
        path_a = tmp_path / "a.pdf"
        path_b = tmp_path / "b.pdf"
        path_a.write_bytes(_make_pdf(1))
        path_b.write_bytes(_make_pdf(1))
        call_count = 0

        class Client:
            async def chat(self, model, messages, response_format=None):
                nonlocal call_count
                call_count += 1
                file_num = (call_count - 1) // 2 + 1
                return {
                    "message": {"content": f'{{"rows": [{{"file": {file_num}}}]}}'},
                    "usage": {},
                }

        source = LLMPagesSource(
            files=(
                FileInput(path=path_a, format="pdf"),
                FileInput(path=path_b, format="pdf"),
            ),
        )
        asyncio.run(ingest_llm_pages(conn, source, "mf_t", client=Client()))
        rows = conn.execute("SELECT file FROM mf_t ORDER BY _row_id").fetchall()
        assert rows == [(1,), (2,)]
        assert call_count == 4

    def test_ingest_llm_pages_page_header_format(self, conn, tmp_path: Path):
        """Verify the page header sent to the LLM."""
        path = tmp_path / "report.pdf"
        path.write_bytes(_make_pdf(2))
        captured_prompts = []

        class Client:
            async def chat(self, model, messages, response_format=None):
                text_parts = [
                    c["text"] for c in messages[0]["content"] if c.get("type") == "text"
                ]
                captured_prompts.append(text_parts[0] if text_parts else "")
                return {
                    "message": {"content": '{"rows": [{"x": 1}]}'},
                    "usage": {},
                }

        source = LLMPagesSource(
            files=(FileInput(path=path, format="pdf"),),
            prompt="Extract rows",
        )
        asyncio.run(ingest_llm_pages(conn, source, "hdr_t", client=Client()))
        # 4 calls: 2 passes * 2 pages
        assert len(captured_prompts) == 4
        assert captured_prompts[0].startswith(
            "=== page 1 of 2 from file report.pdf ==="
        )
        assert "Extract rows" in captured_prompts[0]
        assert captured_prompts[2].startswith(
            "=== page 2 of 2 from file report.pdf ==="
        )

    def test_ingest_llm_pages_no_client(self, conn):
        """LLMPagesSource without client raises RuntimeError."""
        source = LLMPagesSource(files=(FileInput(path=Path("x.pdf"), format="pdf"),))
        with pytest.raises(RuntimeError, match="requires an OpenRouterClient"):
            asyncio.run(ingest_llm_pages(conn, source, "t", client=None))
